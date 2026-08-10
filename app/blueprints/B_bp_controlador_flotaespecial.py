# app/blueprints/B_bp_controlador_flotaespecial.py
import os
from flask import Blueprint, render_template, session, redirect, url_for, request, flash, Response
from app import mysql, bcrypt
from app.utils import login_required_custom
from functools import wraps
import MySQLdb.cursors

bp_controlador_flotaespecial = Blueprint('controlador_flotaespecial', __name__, url_prefix='/gestor_flotaespecial')

def controlador_flotaespecial_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        perfil = str(session.get('perfil', '')).strip().lower()
        tipo_empresa = str(session.get('tipo_empresa', '')).strip().lower()
        
        if perfil not in ['controlador_flotaespecial', 'webmaster'] and 'webmaster' not in tipo_empresa:
            flash('Acceso denegado: Se requiere perfil de Controlador de Transporte Especial para ingresar a este módulo.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# =========================================================
# DASHBOARD PRINCIPAL
# =========================================================
@bp_controlador_flotaespecial.route('/dashboard')
@login_required_custom
@controlador_flotaespecial_required
def dashboard_controlador():
    """
    Dashboard principal del Controlador de Flota Especial.
    """
    return render_template(
        'B_modulo_controlador_flotaespecial.html',
        nit=session.get('nit'),
        empresa=session.get('empresa'),
        nombre=session.get('nombre'),
        active_module='dashboard'
    )

# =========================================================
# GESTIÓN DE VEHÍCULOS
# =========================================================
@bp_controlador_flotaespecial.route('/vehiculos', methods=['GET', 'POST'])
@login_required_custom
@controlador_flotaespecial_required
def gestion_vehiculos():
    empresa_id = session.get('empresa_id')
    empresa_nombre = session.get('empresa')

    if request.method == 'POST':
        accion = request.form.get('accion')
        
        if accion == 'crear':
            placa = str(request.form.get('placa', '')).upper().strip()
            tipo = request.form.get('tipo', '').strip()
            caja_de_carga = request.form.get('caja_de_carga', '').strip()
            referencia = request.form.get('referencia', '').strip()
            peso_vacio = request.form.get('peso_vacio', 0)
            capacidad = request.form.get('capacidad', 0)
            propiedad = request.form.get('propiedad', 'Propio').strip()
            
            if placa and caja_de_carga and tipo:
                cur = mysql.connection.cursor()
                try:
                    cur.execute("""
                        INSERT INTO vehiculos (empresa, id_empresa, placa, tipo, caja_de_carga, referencia, peso_vacio, `capacidad (kg)`, propiedad) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (empresa_nombre, empresa_id, placa, tipo, caja_de_carga, referencia, peso_vacio, capacidad, propiedad))
                    mysql.connection.commit()
                    flash(f"Vehículo con placa {placa} registrado correctamente.", "success")
                except Exception as e:
                    flash(f"Error al registrar vehículo: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'editar':
            vehiculo_id = request.form.get('vehiculo_id')
            placa = str(request.form.get('placa', '')).upper().strip()
            tipo = request.form.get('tipo', '').strip()
            caja_de_carga = request.form.get('caja_de_carga', '').strip()
            referencia = request.form.get('referencia', '').strip()
            peso_vacio = request.form.get('peso_vacio', 0)
            capacidad = request.form.get('capacidad', 0)
            propiedad = request.form.get('propiedad', 'Propio').strip()
            
            if vehiculo_id and placa and caja_de_carga and tipo:
                cur = mysql.connection.cursor()
                try:
                    cur.execute("""
                        UPDATE vehiculos 
                        SET placa = %s, tipo = %s, caja_de_carga = %s, referencia = %s, peso_vacio = %s, `capacidad (kg)` = %s, propiedad = %s
                        WHERE id = %s AND id_empresa = %s
                    """, (placa, tipo, caja_de_carga, referencia, peso_vacio, capacidad, propiedad, vehiculo_id, empresa_id))
                    mysql.connection.commit()
                    flash(f"Vehículo {placa} actualizado correctamente.", "success")
                except Exception as e:
                    flash(f"Error al actualizar vehículo: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'eliminar':
            vehiculo_id = request.form.get('vehiculo_id')
            cur = mysql.connection.cursor()
            try:
                cur.execute("DELETE FROM vehiculos WHERE id = %s AND id_empresa = %s", (vehiculo_id, empresa_id))
                mysql.connection.commit()
                flash("Vehículo eliminado de la base de datos.", "success")
            except Exception as e:
                flash("Error al eliminar vehículo.", "danger")
            finally:
                cur.close()

        return redirect(url_for('controlador_flotaespecial.gestion_vehiculos'))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT * FROM vehiculos WHERE id_empresa = %s ORDER BY id DESC", (empresa_id,))
    vehiculos_db = cur.fetchall()
    cur.close()

    return render_template(
        'B_modulo_controlador_flotaespecial.html',
        nit=session.get('nit'),
        empresa=session.get('empresa'),
        nombre=session.get('nombre'),
        active_module='vehiculos', 
        vehiculos=vehiculos_db
    )

@bp_controlador_flotaespecial.route('/vehiculos/plantilla', methods=['GET'])
@login_required_custom
@controlador_flotaespecial_required
def descargar_plantilla_vehiculos():
    """Genera y descarga la plantilla Excel (.xlsx) para carga masiva de Vehículos"""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    
    # 1. Pestaña Principal (Plantilla)
    ws1 = wb.active
    ws1.title = "Plantilla"
    encabezados = ['Placa', 'Tipo de Vehiculo', 'Adaptacion / Clase', 'Marca / Referencia', 'Propiedad', 'Peso Vacio (Kg)', 'Capacidad (Pasajeros)']
    ws1.append(encabezados)
    
    # Ajustar ancho de columnas para mejor visibilidad
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws1.column_dimensions[col].width = 20

    # 2. Pestaña Guía y Ejemplos
    ws2 = wb.create_sheet(title="Guía y Ejemplos")
    ws2.append(["INSTRUCCIONES PARA LA CARGA MASIVA"])
    ws2.append(["1. Llene los datos de los vehículos ÚNICAMENTE en la pestaña 'Plantilla'."])
    ws2.append(["2. No modifique ni elimine los encabezados de la primera fila."])
    ws2.append(["3. Copie y pegue EXACTAMENTE los valores permitidos para las listas desplegables."])
    ws2.append([])
    ws2.append(["COLUMNA", "VALORES PERMITIDOS EXACTOS (Copiar y Pegar)"])
    ws2.append(["Tipo de Vehiculo", "van, buseta, bus, ambulancia, sedan, camioneta"])
    ws2.append(["Adaptacion / Clase", "Transporte de Pacientes, Transporte Escolar, Transporte Empresarial, Turismo, Carga Especial, Otro"])
    ws2.append(["Propiedad", "Propio, Tercero"])
    ws2.append([])
    ws2.append(["EJEMPLO DE LLENADO CORRECTO EN LA PESTAÑA 'Plantilla':"])
    ws2.append(['AAA123', 'sedan', 'Transporte Empresarial', 'Chevrolet', 'Propio', 1200, 4])
    ws2.append(['BBB456', 'van', 'Transporte Escolar', 'Renault', 'Tercero', 1500, 15])
    
    for col in ['A', 'B']:
        ws2.column_dimensions[col].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=plantilla_vehiculos.xlsx"}
    )

@bp_controlador_flotaespecial.route('/vehiculos/carga_masiva', methods=['POST'])
@login_required_custom
@controlador_flotaespecial_required
def carga_masiva_vehiculos():
    """Procesa el archivo Excel (.xlsx) subido para carga masiva de Vehículos"""
    import openpyxl

    empresa_id = session.get('empresa_id')
    empresa_nombre = session.get('empresa')
    
    if 'archivo_excel' not in request.files:
        flash('No se subió ningún archivo.', 'danger')
        return redirect(url_for('controlador_flotaespecial.gestion_vehiculos'))
        
    file = request.files['archivo_excel']
    if file.filename == '':
        flash('Ningún archivo seleccionado.', 'danger')
        return redirect(url_for('controlador_flotaespecial.gestion_vehiculos'))
        
    if not file.filename.endswith('.xlsx'):
        flash('El archivo debe ser un formato Excel (.xlsx) válido.', 'danger')
        return redirect(url_for('controlador_flotaespecial.gestion_vehiculos'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        # Seleccionar la pestaña de la plantilla (evitar procesar la hoja de instrucciones)
        if 'Plantilla' in wb.sheetnames:
            ws = wb['Plantilla']
        else:
            ws = wb.active # Fallback
            
        cur = mysql.connection.cursor()
        registros_exitosos = 0
        
        # Iterar filas (values_only=True extrae solo los datos, no las celdas como objeto)
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1: continue # Saltar fila de encabezados
            if not any(row): continue # Saltar filas totalmente vacías
            
            # Aseguramos que la fila tenga al menos 7 columnas leyendo de forma segura
            placa = str(row[0]).upper().strip() if len(row) > 0 and row[0] is not None else ""
            tipo = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            caja_de_carga = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            referencia = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            propiedad = str(row[4]).strip() if len(row) > 4 and row[4] is not None else "Propio"
            
            try:
                peso_vacio = int(row[5]) if len(row) > 5 and row[5] is not None else 0
            except (ValueError, TypeError):
                peso_vacio = 0
                
            try:
                capacidad = int(row[6]) if len(row) > 6 and row[6] is not None else 0
            except (ValueError, TypeError):
                capacidad = 0
            
            # Ignorar si no hay placa
            if placa and not placa.startswith('EJ.'):
                try:
                    cur.execute("""
                        INSERT INTO vehiculos (empresa, id_empresa, placa, tipo, caja_de_carga, referencia, peso_vacio, `capacidad (kg)`, propiedad) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (empresa_nombre, empresa_id, placa, tipo, caja_de_carga, referencia, peso_vacio, capacidad, propiedad))
                    registros_exitosos += 1
                except Exception as inner_e:
                    # Falla silenciosa si la placa ya existe para que el ciclo continúe
                    print(f"Error insertando vehículo {placa}: {inner_e}")
                    pass

        mysql.connection.commit()
        cur.close()
        flash(f"Carga masiva completada: Se registraron {registros_exitosos} vehículos exitosamente.", "success")
    except Exception as e:
        flash(f"Error procesando el archivo Excel: {str(e)}", "danger")

    return redirect(url_for('controlador_flotaespecial.gestion_vehiculos'))

# =========================================================
# GESTIÓN DE OPERADORES
# =========================================================
@bp_controlador_flotaespecial.route('/operadores', methods=['GET', 'POST'])
@login_required_custom
@controlador_flotaespecial_required
def gestion_operadores():
    empresa_id = session.get('empresa_id')
    empresa_nombre = session.get('empresa')

    if request.method == 'POST':
        accion = request.form.get('accion')
        
        if accion == 'crear':
            nombre = request.form.get('nombre', '').strip()
            cedula = request.form.get('cedula', '').strip()
            perfil = request.form.get('perfil', '').strip()
            
            if perfil == 'operador_flotaespecial':
                password = request.form.get('password', '').strip()
                if not password:
                    flash("El operador requiere una contraseña de acceso.", "danger")
                    return redirect(url_for('controlador_flotaespecial.gestion_operadores'))
                hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
            else:
                hashed_pw = bcrypt.generate_password_hash(os.urandom(12).hex()).decode('utf-8')

            if nombre and cedula and perfil:
                cur = mysql.connection.cursor()
                try:
                    cur.execute("SELECT id FROM usuarios WHERE cedula = %s", (cedula,))
                    if cur.fetchone():
                        flash(f"La identificación {cedula} ya está registrada.", "danger")
                    else:
                        cur.execute("""
                            INSERT INTO usuarios (nombre, cedula, password, perfil, empresa, empresa_id) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (nombre, cedula, hashed_pw, perfil, empresa_nombre, empresa_id))
                        mysql.connection.commit()
                        flash(f"Personal registrado exitosamente: {nombre}.", "success")
                except Exception as e:
                    flash(f"Error al registrar: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'editar':
            operador_id = request.form.get('operador_id')
            nombre = request.form.get('nombre', '').strip()
            cedula = request.form.get('cedula', '').strip()
            perfil = request.form.get('perfil', '').strip()
            
            if operador_id and nombre and cedula and perfil:
                cur = mysql.connection.cursor()
                try:
                    if perfil == 'operador_flotaespecial':
                        password = request.form.get('password', '').strip()
                        if password:
                            hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
                            cur.execute("""
                                UPDATE usuarios 
                                SET nombre = %s, cedula = %s, perfil = %s, password = %s
                                WHERE id = %s AND empresa_id = %s
                            """, (nombre, cedula, perfil, hashed_pw, operador_id, empresa_id))
                        else:
                            cur.execute("""
                                UPDATE usuarios 
                                SET nombre = %s, cedula = %s, perfil = %s
                                WHERE id = %s AND empresa_id = %s
                            """, (nombre, cedula, perfil, operador_id, empresa_id))
                    else:
                        cur.execute("""
                            UPDATE usuarios 
                            SET nombre = %s, cedula = %s, perfil = %s
                            WHERE id = %s AND empresa_id = %s
                        """, (nombre, cedula, perfil, operador_id, empresa_id))
                        
                    mysql.connection.commit()
                    flash(f"Registro actualizado correctamente.", "success")
                except Exception as e:
                    flash(f"Error al actualizar: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'eliminar':
            operador_id = request.form.get('operador_id')
            cur = mysql.connection.cursor()
            try:
                cur.execute("DELETE FROM usuarios WHERE id = %s AND empresa_id = %s", (operador_id, empresa_id))
                mysql.connection.commit()
                flash("Registro eliminado permanentemente.", "success")
            except Exception as e:
                flash("Error al eliminar.", "danger")
            finally:
                cur.close()

        return redirect(url_for('controlador_flotaespecial.gestion_operadores'))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    # Filtrar solo el personal de transporte especial
    cur.execute("""
        SELECT id, nombre, cedula, perfil 
        FROM usuarios 
        WHERE empresa_id = %s AND perfil IN ('operador_flotaespecial', 'auxiliar_transporte_especial')
        ORDER BY nombre ASC
    """, (empresa_id,))
    operadores_db = cur.fetchall()
    cur.close()

    return render_template(
        'B_modulo_controlador_flotaespecial.html',
        nit=session.get('nit'),
        empresa=session.get('empresa'),
        nombre=session.get('nombre'),
        active_module='operadores', 
        operadores=operadores_db
    )

# =========================================================
# GESTIÓN EPS Y AUTORIZACIONES
# =========================================================
@bp_controlador_flotaespecial.route('/gestion_eps', methods=['GET', 'POST'])
@login_required_custom
@controlador_flotaespecial_required
def gestion_eps():
    empresa_id = session.get('empresa_id')
    empresa_nombre = session.get('empresa')

    if request.method == 'POST':
        accion = request.form.get('accion')
        
        if accion == 'crear':
            cliente_empresa = request.form.get('cliente_empresa', '').strip()
            id_cliente_empresa = request.form.get('id_cliente_empresa', '').strip()
            
            if cliente_empresa and id_cliente_empresa:
                cur = mysql.connection.cursor()
                try:
                    cur.execute("""
                        INSERT INTO clientes_empresa (empresa, id_empresa, cliente_empresa, id_cliente_empresa) 
                        VALUES (%s, %s, %s, %s)
                    """, (empresa_nombre, empresa_id, cliente_empresa, id_cliente_empresa))
                    mysql.connection.commit()
                    flash(f"EPS '{cliente_empresa}' registrada correctamente.", "success")
                except Exception as e:
                    flash(f"Error al registrar la EPS: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'editar':
            eps_id = request.form.get('eps_id')
            cliente_empresa = request.form.get('cliente_empresa', '').strip()
            id_cliente_empresa = request.form.get('id_cliente_empresa', '').strip()
            
            if eps_id and cliente_empresa and id_cliente_empresa:
                cur = mysql.connection.cursor()
                try:
                    cur.execute("""
                        UPDATE clientes_empresa 
                        SET cliente_empresa = %s, id_cliente_empresa = %s
                        WHERE id = %s AND id_empresa = %s
                    """, (cliente_empresa, id_cliente_empresa, eps_id, empresa_id))
                    mysql.connection.commit()
                    flash(f"EPS '{cliente_empresa}' actualizada correctamente.", "success")
                except Exception as e:
                    flash(f"Error al actualizar la EPS: {str(e)}", "danger")
                finally:
                    cur.close()

        elif accion == 'eliminar':
            eps_id = request.form.get('eps_id')
            cur = mysql.connection.cursor()
            try:
                cur.execute("DELETE FROM clientes_empresa WHERE id = %s AND id_empresa = %s", (eps_id, empresa_id))
                mysql.connection.commit()
                flash("EPS eliminada del sistema.", "success")
            except Exception as e:
                flash("Error al eliminar la EPS.", "danger")
            finally:
                cur.close()

        return redirect(url_for('controlador_flotaespecial.gestion_eps'))

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT * FROM clientes_empresa WHERE id_empresa = %s ORDER BY id DESC", (empresa_id,))
    eps_db = cur.fetchall()
    cur.close()

    return render_template(
        'B_modulo_controlador_flotaespecial.html',
        nit=session.get('nit'),
        empresa=session.get('empresa'),
        nombre=session.get('nombre'),
        active_module='gestion_eps', 
        lista_eps=eps_db
    )

@bp_controlador_flotaespecial.route('/gestion_autorizaciones', methods=['GET'])
@login_required_custom
@controlador_flotaespecial_required
def gestion_autorizaciones():
    return render_template(
        'B_modulo_controlador_flotaespecial.html',
        nit=session.get('nit'),
        empresa=session.get('empresa'),
        nombre=session.get('nombre'),
        active_module='gestion_autorizaciones'
    )